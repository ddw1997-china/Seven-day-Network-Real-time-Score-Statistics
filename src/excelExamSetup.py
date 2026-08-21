import sys
import re
import os
import pandas as pd
from collections import deque
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from PyQt6.QtWidgets import (
    QApplication,  QWidget, QVBoxLayout, QHBoxLayout, QLabel,
    QPushButton, QFileDialog, QRadioButton, QGroupBox, QMessageBox, QLineEdit
)
from PyQt6.QtCore import Qt


def clean_sheet_name(name):
    if not isinstance(name, str):
        print(f"Warning: class name '{name}' is not a string. Converting to string.")
        name = str(name)
    name = re.sub(r'[\\/*?:"<>|]', '', name)
    return name[:31]


def check_any_in(listA, listB):
    for element in listA:
        if element in listB:
            return element
    return None


class ExamAllocationApp(QWidget):
    def __init__(self):
        super().__init__()
        # 主部件和布局
        layout = QVBoxLayout()
        self.setLayout(layout)
        self.setObjectName('examSetup')
        # 标题
        title_label = QLabel("根据成绩分配考场及座位")
        title_label.setStyleSheet("font-size: 18px; font-weight: bold;")
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title_label)

        # 规则选择部分
        self.rule_group = QGroupBox("选择分配规则")
        self.rule_group.setObjectName("rule_group")
        rule_layout = QVBoxLayout()
        rule_layout.setSpacing(20)
        rule_layout.setContentsMargins(50, 20, 0, 10)

        self.rule0_radio = QRadioButton("规则一: 分班排序再依班轮询筛选")
        self.rule1_radio = QRadioButton("规则二: 分班排序再集中名次筛选")
        self.rule2_radio = QRadioButton("规则三: 全部集中排序再依次筛选")

        self.rule0_radio.setChecked(True)

        rule_layout.addWidget(self.rule0_radio)
        rule_layout.addWidget(self.rule1_radio)
        rule_layout.addWidget(self.rule2_radio)
        self.rule_group.setLayout(rule_layout)
        layout.addWidget(self.rule_group)

        # 规则说明部分
        self.rule_description = QLabel()
        self.rule_description.setObjectName("rule_description")
        self.rule_description.setContentsMargins(50,20,10,20)
        self.rule_description.setWordWrap(True)
        self.update_rule_description()
        layout.addWidget(self.rule_description)

        # 连接单选按钮的信号
        self.rule0_radio.toggled.connect(self.update_rule_description)
        self.rule1_radio.toggled.connect(self.update_rule_description)
        self.rule2_radio.toggled.connect(self.update_rule_description)

        # 考点代码输入
        code_layout = QHBoxLayout()
        code_label = QLabel("考点代码:")
        self.code_input = QLineEdit()
        self.code_input.setText("711")
        self.code_input.setMaximumWidth(100)
        code_layout.addStretch()
        code_layout.addWidget(code_label)
        code_layout.addWidget(self.code_input)
        layout.addLayout(code_layout)

        # 按钮部分
        button_layout = QHBoxLayout()
        self.select_capacity_button = QPushButton("选择考场人数设置文件")
        self.select_capacity_button.setFixedWidth(150)
        self.select_capacity_button.clicked.connect(self.select_capacity_file)
        self.select_students_button = QPushButton("选择学生信息文件")
        self.select_students_button.clicked.connect(self.select_students_file)
        self.select_students_button.setFixedWidth(120)
        self.select_students_button.setEnabled(False)

        button_layout.addStretch()
        button_layout.addWidget(self.select_capacity_button)
        button_layout.addStretch()
        button_layout.addWidget(self.select_students_button)
        button_layout.addStretch()
        layout.addLayout(button_layout)

        # 状态标签
        self.status_label = QLabel("请先选择考场人数设置文件")
        self.status_label.setObjectName("status_label")
        self.status_label.setWordWrap(True)
        layout.addWidget(self.status_label)

        # 添加伸缩因子
        layout.addStretch()

        # 文件路径存储
        self.capacity_file_path = None
        self.students_file_path = None

    def update_rule_description(self):
        """更新规则说明"""
        if self.rule0_radio.isChecked():
            self.rule_description.setText(
                "规则一说明：\n\n"
                "分班每轮依次筛选出名次最靠前的。\n"
                "特点：每轮依次从各班取最靠前的一名，避免同一轮连续抽取同班学生。\n"
                "缺点：如果班级差距较大，同一考场可能学生分差较大。\n"
                "适用场景：最大限度分散同班学生，同一考场尽量安排不同班级。"
            )
        elif self.rule1_radio.isChecked():
            self.rule_description.setText(
                "规则二说明：\n\n"
                "按班级名次集中筛选。\n"
                "特点：先取各班第1名，再取各班第2名，依次类推。\n"
                "缺点：如果班级差距较大，同一考场可能学生分差较大。\n"
                "适用场景：最大限度分散同班学生，需要按班级依次均匀分布的考试。"
            )
        elif self.rule2_radio.isChecked():
            self.rule_description.setText(
                "规则三说明：\n\n"
                "集中全部按总分排序后再分配。\n"
                "特点：完全按成绩高低排序，同班学生可能连续坐在一起。\n"
                "缺点：如果班级差距较大，同一考场可能同班学生会连续抽取。\n"
                "适用场景：完全按成绩排名的考试。"
            )

    def select_capacity_file(self):
        """选择考场人数设置文件"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择考场设置文件", "", "Excel文件 (*.xlsx *.xls)"
        )

        if file_path:
            self.capacity_file_path = file_path
            self.exam_room_capacity = self.read_exam_room_capacity(self.capacity_file_path)
            if not self.exam_room_capacity:
                # self.creat_dialog_window(self,"capacity_file",
                #                          "警告","请选择正确的Excel配置文件",
                #                          QMessageBox.StandardButton.Close,
                #                          QMessageBox.Icon.Critical)
                return
            self.status_label.setText(f"已选择考场设置文件: {file_path}\n\n请选择学生信息文件")
            self.select_students_button.setEnabled(True)

    def select_students_file(self):
        """选择学生信息文件"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择学生信息文件", "", "Excel文件 (*.xlsx *.xls)"
        )

        if file_path:
            self.students_file_path = file_path
            self.status_label.setText(
                f"考场设置文件: {os.path.basename(self.capacity_file_path)}\n"
                f"学生信息文件: {os.path.basename(file_path)}\n"
                "正在处理..."
            )
            result = self.creat_dialog_window(self,"info_file",
                                     "通知","请确认文件，按Yes开始！",
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                                     QMessageBox.Icon.Information)
            if result != QMessageBox.StandardButton.Yes:
                self.students_file_path = ""
                self.capacity_file_path = ""
                self.status_label.setText(
                    "已重置，先选择考场设置文件"
                )
                self.select_students_button.setEnabled(False)
                return
            self.process_files()

    def read_exam_room_capacity(self, file_path):
        """读取考场容量设置"""
        try:
            capacity_df = pd.read_excel(file_path)
            room = check_any_in(['考场号', '考场', '考室号', '考室'], capacity_df.columns)
            volume = check_any_in(['容量', '人数', '数量', '大小'], capacity_df.columns)

            if not room or not volume:
                QMessageBox.critical(self, "错误", "考场设置文件必须包含 '考场' 和 '人数' 列")
                return None

            return capacity_df.set_index(capacity_df[f'{room}'].astype(int))[f'{volume}'].to_dict()
        except FileNotFoundError:
            QMessageBox.critical(self, "错误", "考场设置文件未找到，请检查文件路径")
            return None
        except Exception as e:
            QMessageBox.critical(self, "错误", f"读取考场设置文件时出错: {str(e)}")
            print(f"详细错误信息: {e}")
            return None

    def assign_exam_rooms(self, students_file_path, exam_room_capacity, examcode):
        """分配考场座位"""
        try:
            # 确定选中的规则
            if self.rule0_radio.isChecked():
                rule = 0
            elif self.rule1_radio.isChecked():
                rule = 1
            else:
                rule = 2

            # 存储每个工作表的学生信息列表
            all_students = []
            tables = []
            xls = pd.ExcelFile(students_file_path)
            sheet_names = xls.sheet_names

            for ind in range(len(sheet_names)):
                sheet_name = sheet_names[ind]
                df = pd.read_excel(xls, sheet_name)
                name_ = check_any_in(['姓名', '考生', '考生姓名', '学生', '学生姓名', '名字'], df.columns)
                score_ = check_any_in(['分数', '成绩', '总分', '总成绩', '总分数', '综合成绩'], df.columns)
                class_ = check_any_in(['班级', '班名'], df.columns)

                if not name_ or not score_ or not class_:
                    QMessageBox.critical(
                        self,
                        "错误",
                        f"工作表 '{sheet_name}' 必须包含 '姓名'、'成绩'、'班级' 列"
                    )
                    if ind == len(sheet_names) - 1:
                        self.status_label.setText(
                            f"考场设置文件: {os.path.basename(self.capacity_file_path)}\n"
                            f"学生信息文件: 请重选学生信息Excel文件\n"
                            "等候处理..."
                        )
                        return None
                    else:
                        continue

                df = df.sort_values(by=f'{score_}', ascending=False)
                df['总分'] = df[score_].astype(float)
                df['班级名次'] = df[score_].rank(ascending=False, method='min')
                df['班级'] = df[class_].astype(str)
                df['姓名'] = df[name_].astype(str)

                tables.append(deque(df.to_dict('records')))
                all_students.append(df)

            # 第一种编排方案：每班数据分别提取头名
            if rule == 0:
                allocation = {}
                table_index = 0

                for exam_room_index, capacity in exam_room_capacity.items():
                    examroom = []
                    seat_number = 1

                    while len(examroom) < capacity:
                        allocated = False
                        table_index %= len(tables)
                        table = tables[table_index]

                        if len(table) == 0:
                            table_index += 1
                            continue

                        student = table.popleft()
                        name = student['姓名']
                        class_name = student['班级']
                        original_rank = student['班级名次']

                        if original_rank is None:
                            original_rank = 0

                        seat_str = str(seat_number).zfill(2)
                        exam_room_index_str = str(exam_room_index).zfill(2)
                        exam_number = f"{examcode}{exam_room_index_str}{seat_str}"

                        examroom.append((name, seat_str, class_name, original_rank, exam_room_index_str, exam_number))
                        allocated = True
                        table_index += 1
                        seat_number += 1

                        if not allocated and len(examroom) < capacity:
                            raise RuntimeError(
                                f"考场{exam_room_index_str}需要{capacity}人，但可用学生只有{len(examroom)}人")

                    allocation[f"考场{exam_room_index_str}"] = examroom

                return allocation

            # 第二种和第三种编排方案
            all_students_df = pd.concat(all_students, ignore_index=True)

            if rule == 1:
                all_students_df = all_students_df.sort_values(by=['班级名次', '总分'], ascending=[True, False])
            elif rule == 2:
                all_students_df = all_students_df.sort_values(by=['总分', '班级名次'], ascending=[False, True])

            exam_rooms = {}
            start_index = 0

            for exam_room_number, capacity in exam_room_capacity.items():
                end_index = start_index + capacity
                students_in_room = all_students_df.iloc[start_index:end_index]
                start_index = end_index
                exam_rooms[exam_room_number] = []
                seat_number = 1

                for index, student_row in students_in_room.iterrows():
                    name = student_row['姓名']
                    class_name = student_row['班级']
                    original_rank = student_row['班级名次']

                    if original_rank is None:
                        original_rank = 0

                    seat_str = str(seat_number).zfill(2)
                    exam_room_number_str = str(exam_room_number).zfill(2)
                    exam_number = f"{examcode}{exam_room_number_str}{seat_str}"

                    exam_rooms[exam_room_number].append(
                        (name, seat_str, class_name, original_rank, exam_room_number_str, exam_number))
                    seat_number += 1

            return exam_rooms

        except Exception as e:
            QMessageBox.critical(self, "错误", f"处理学生信息时出错: {str(e)}")
            print(f"详细错误信息: {e}")
            return None

    def save_results(self, result_df):
        """保存结果到文件"""
        save_path, _ = QFileDialog.getSaveFileName(
            self, "保存结果", "", "Excel文件 (*.xlsx)"
        )

        if not save_path:
            self.status_label.setText("已取消保存，请先选择Excel配置文件......")
            self.select_students_button.setEnabled(False)
            self.students_file_path = ""
            self.capacity_file_path = ""
            return

        try:
            wb = Workbook()

            # 按照班级分割
            for class_name, class_df in result_df.groupby('原班级'):
                clean_class_name = clean_sheet_name(class_name)
                class_sheet = wb.create_sheet(title=clean_class_name)
                class_sheet.append(list(class_df.columns))
                rows = dataframe_to_rows(class_df, index=False, header=False)
                for row in rows:
                    class_sheet.append(row)

            # 按照考场号分割
            for exam_room_number, exam_room_df in result_df.groupby('考场号'):
                clean_exam_room_name = clean_sheet_name(f"考场_{exam_room_number}")
                exam_room_df_sheet = wb.create_sheet(title=clean_exam_room_name)
                exam_room_df_sheet.append(list(exam_room_df.columns))
                rows = dataframe_to_rows(exam_room_df, index=False, header=False)
                for row in rows:
                    exam_room_df_sheet.append(row)

            # 考场汇总表
            exam_room_sheet = wb.create_sheet(title="考场汇总")
            exam_room_sheet.append(list(result_df.columns))
            for exam_room_number, exam_room_df in result_df.groupby('考场号'):
                rows = dataframe_to_rows(exam_room_df, index=False, header=False)
                for row in rows:
                    exam_room_sheet.append(row)

            wb.remove(wb['Sheet'])
            wb.save(save_path)

            # 尝试打开保存的文件
            try:
                os.startfile(save_path)
            except:
                pass

            QMessageBox.information(self, "完成", f"结果已保存到: {save_path}")
            self.status_label.setText("处理完成，结果已保存")

        except Exception as e:
            QMessageBox.critical(self, "错误", f"保存文件时出错: {str(e)}")
            print(f"详细错误信息: {e}")

    def process_files(self):
        """处理选中的文件"""
        if not self.students_file_path:
            return

        examcode = self.code_input.text().strip()
        if not examcode:
            examcode = "711"

        exam_rooms = self.assign_exam_rooms(self.students_file_path, self.exam_room_capacity, examcode)
        if not exam_rooms:
            return

        # 准备结果数据
        result_list = []
        for room, students in exam_rooms.items():
            for name, seat_number, class_name, original_rank, exam_room_number, exam_number in students:
                result_list.append({
                    '考号': exam_number,
                    '考场号': exam_room_number,
                    '姓名': name,
                    '座位号': seat_number,
                    '原班级': class_name,
                    '班级名次': original_rank
                })

        result_df = pd.DataFrame(result_list)
        result_df = result_df.fillna("")

        self.save_results(result_df)

    def creat_dialog_window(self,_parent,_objectname,_title,_text,_buttons,_icon):
        dialog = QMessageBox(_parent)
        dialog.setObjectName(_objectname)
        dialog.setWindowTitle(_title)
        dialog.setText(_text)

        # 关键设置
        dialog.setWindowFlags(
            dialog.windowFlags() |
            Qt.WindowType.WindowStaysOnTopHint  # 置顶
        )
        dialog.setStandardButtons(_buttons)
        dialog.setIcon(_icon)
        dialog.setModal(True)  # 设为模态（阻塞父窗口）
        result = dialog.exec()  # 显示弹窗（阻塞）
        return result
if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = ExamAllocationApp()
    window.show()
    sys.exit(app.exec())