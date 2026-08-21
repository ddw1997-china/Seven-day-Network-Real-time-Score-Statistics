import sys
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout,
                             QHBoxLayout, QPushButton, QTableWidget,
                             QTableWidgetItem, QFileDialog, QMessageBox,
                             QStyledItemDelegate, QComboBox)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QColor


class ComboDelegate(QStyledItemDelegate):
    def __init__(self, items):
        super().__init__()
        self.items = [""] + items  # 添加空选项

    def createEditor(self, parent, option, index):
        editor = QComboBox(parent)
        editor.addItems(self.items)
        return editor

    def setEditorData(self, editor, index):
        value = index.data(Qt.ItemDataRole.DisplayRole) or ""
        editor.setCurrentText(value)

    def setModelData(self, editor, model, index):
        value = editor.currentText()
        model.setData(index, value, Qt.ItemDataRole.EditRole)

    def editorEvent(self, event, model, option, index):
        if event.type() == event.Type.MouseButtonPress:
            editor = self.createEditor(None, option, index)
            editor.showPopup()
            return True
        return super().editorEvent(event, model, option, index)


class ExcelDataProcessor(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()
        self.workbook = None
        self.summary_sheet = None

    def initUI(self):
        layout = QVBoxLayout()
        self.setLayout(layout)
        self.setObjectName("sum_table")

        button_layout = QHBoxLayout()
        self.load_btn = QPushButton('加载Excel文件')
        self.load_btn.clicked.connect(self.load_excel)
        button_layout.addWidget(self.load_btn)

        self.process_btn = QPushButton('处理数据')
        self.process_btn.clicked.connect(self.process_data)
        self.process_btn.setEnabled(False)
        button_layout.addWidget(self.process_btn)

        self.save_btn = QPushButton('保存结果')
        self.save_btn.clicked.connect(self.save_results)
        self.save_btn.setEnabled(False)
        button_layout.addWidget(self.save_btn)

        layout.addLayout(button_layout)

        self.table = QTableWidget()
        self.table.setStyleSheet("""
            QTableWidget {
                gridline-color: #c0c0c0;
                font-size: 12px;
            }
            QTableWidget::item[status="check"] {
                background-color: #ffcccc;
            }
            QTableWidget::item[status="circle"] {
                background-color: #ccffcc;
            }
        """)
        layout.addWidget(self.table)

    def load_excel(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择Excel文件", "", "Excel文件 (*.xlsx *.xls)"
        )

        if file_path:
            try:
                self.workbook = openpyxl.load_workbook(file_path)
                self.process_btn.setEnabled(True)
                QMessageBox.information(self, "成功", "Excel文件加载成功！")
            except Exception as e:
                QMessageBox.critical(self, "错误", f"加载文件失败: {str(e)}")

    def process_data(self):
        if not self.workbook:
            QMessageBox.warning(self, "警告", "请先加载Excel文件")
            return

        try:
            if "总表" in self.workbook.sheetnames:
                self.summary_sheet = self.workbook["总表"]
                self.workbook.remove(self.summary_sheet)

            self.summary_sheet = self.workbook.create_sheet("总表", 0)

            sheet_names = [name for name in self.workbook.sheetnames if name != "总表"]
            data_dict = {}

            for sheet_name in sheet_names:
                sheet = self.workbook[sheet_name]
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value and str(cell.value).strip() == "姓名":
                            start_row = cell.row
                            start_col = cell.column
                            end_col = start_col
                            while end_col <= sheet.max_column:
                                if sheet.cell(row=start_row, column=end_col).value is None:
                                    break
                                end_col += 1
                            end_col -= 1

                            end_row = start_row
                            while end_row <= sheet.max_row:
                                if sheet.cell(row=end_row, column=start_col).value is None:
                                    break
                                end_row += 1
                            end_row -= 1

                            if end_row > start_row and end_col >= start_col:
                                for row_idx in range(start_row + 1, end_row + 1):
                                    name = sheet.cell(row=row_idx, column=start_col).value
                                    if name:
                                        try:
                                            value = float(sheet.cell(row=row_idx, column=end_col).value)
                                            if name not in data_dict:
                                                data_dict[name] = {}
                                            current_value = data_dict[name].get(sheet_name, 0)
                                            current_value += value
                                            data_dict[name][sheet_name] = current_value
                                        except (ValueError, TypeError):
                                            continue

            self.generate_summary_structure(data_dict, sheet_names)
            self.display_summary_data(data_dict, sheet_names)
            self.save_btn.setEnabled(True)
            QMessageBox.information(self, "成功", "数据处理完成！")

        except Exception as e:
            QMessageBox.critical(self, "错误", f"处理数据失败: {str(e)}")

    def generate_summary_structure(self, data_dict, sheet_names):
        headers = ["姓名"] + sheet_names + ["总数", "备注"]
        self.summary_sheet.append(headers)

        bold_font = Font(bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col in range(1, len(headers) + 1):
            cell = self.summary_sheet.cell(row=1, column=col)
            cell.font = bold_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

    def display_summary_data(self, data_dict, sheet_names):
        headers = ["姓名"] + sheet_names + ["总数", "备注"]
        row_data = []

        for name, sheet_data in data_dict.items():
            row = [name]
            total = 0

            for sheet in sheet_names:
                value = sheet_data.get(sheet, 0)
                row.append(str(value))
                total += value

            row.append(str(total))
            row.append("")  # 初始化为空备注
            row_data.append(row)

        self.table.setColumnCount(len(headers))
        self.table.setRowCount(len(row_data))
        self.table.setHorizontalHeaderLabels(headers)

        for row_idx, row in enumerate(row_data):
            for col_idx, value in enumerate(row):
                item = QTableWidgetItem(str(value))
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

                if col_idx == len(headers) - 1:
                    item.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsEditable)
                else:
                    item.setFlags(Qt.ItemFlag.ItemIsEnabled)

                self.table.setItem(row_idx, col_idx, item)

        # 设置下拉列表委托
        combo_delegate = ComboDelegate(["√", "○"])
        self.table.setItemDelegateForColumn(len(headers) - 1, combo_delegate)
        self.table.cellChanged.connect(self.handle_cell_changed)
        self.table.resizeColumnsToContents()

    def handle_cell_changed(self, row, col):
        """改进的单元格变更处理"""
        if col != self.table.columnCount() - 1:  # 只处理备注列
            return

        # 防止递归调用
        self.table.blockSignals(True)

        try:
            item = self.table.item(row, col)
            value = item.text() if item else ""

            # 设置整行样式
            for c in range(self.table.columnCount()):
                cell = self.table.item(row, c)
                if cell:
                    if value == "√":
                        cell.setBackground(QColor(255, 200, 200))  # 浅红
                    elif value == "○":
                        cell.setBackground(QColor(200, 255, 200))  # 浅绿
                    else:
                        cell.setBackground(QColor(255, 255, 255))  # 白色
        finally:
            self.table.blockSignals(False)

    def save_results(self):
        if not self.summary_sheet:
            QMessageBox.warning(self, "警告", "没有可保存的数据")
            return

        try:
            # 1. 将表格数据写入Excel
            for row in range(self.table.rowCount()):
                for col in range(self.table.columnCount()):
                    item = self.table.item(row, col)
                    if item:
                        self.summary_sheet.cell(
                            row=row + 2,  # 从第2行开始写入（第1行是标题）
                            column=col + 1,
                            value=item.text()
                        )

            # 2. 添加备注列的下拉列表验证
            notes_col = len(self.summary_sheet[1])  # 获取备注列位置
            notes_col_letter = openpyxl.utils.get_column_letter(notes_col)

            # 创建数据验证规则
            dv = DataValidation(
                type="list",
                formula1='",√,○"',  # 定义下拉选项
                allow_blank=True,  # 允许空值
                showErrorMessage=True,  # 显示错误提示
                errorTitle="无效输入",
                error="请从下拉列表选择√或○"
            )

            # 应用数据验证范围（从第2行到最后一行）
            start_row = 2
            end_row = self.table.rowCount() + 1
            dv.add(f"{notes_col_letter}{start_row}:{notes_col_letter}{end_row}")
            self.summary_sheet.add_data_validation(dv)

            # 3. 添加条件格式（根据备注值设置行颜色）
            self.add_conditional_formatting()

            # 4. 保存文件
            file_path, _ = QFileDialog.getSaveFileName(
                self, "保存Excel文件", "", "Excel文件 (*.xlsx)"
            )

            if file_path:
                if not file_path.endswith('.xlsx'):
                    file_path += '.xlsx'

                self.workbook.save(file_path)
                QMessageBox.information(self, "成功", "文件保存成功！")

        except PermissionError:
            QMessageBox.critical(self, "错误", "文件被其他程序占用，请关闭后重试")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"保存失败: {str(e)}")

    def add_conditional_formatting(self):
        """为备注列添加条件格式"""
        notes_col = len(self.summary_sheet[1])
        notes_col_letter = openpyxl.utils.get_column_letter(notes_col)
        max_row = self.summary_sheet.max_row

        # 红色格式（当备注为√时）
        red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        red_rule = FormulaRule(
            formula=[f'${notes_col_letter}2="√"'],  # 使用绝对列引用
            stopIfTrue=False,
            fill=red_fill
        )
        self.summary_sheet.conditional_formatting.add(
            f"A2:{openpyxl.utils.get_column_letter(notes_col - 1)}{max_row}",
            red_rule
        )

        # 绿色格式（当备注为○时）
        green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        green_rule = FormulaRule(
            formula=[f'${notes_col_letter}2="○"'],
            stopIfTrue=False,
            fill=green_fill
        )
        self.summary_sheet.conditional_formatting.add(
            f"A2:{openpyxl.utils.get_column_letter(notes_col - 1)}{max_row}",
            green_rule
        )
        # 3. 默认白色格式（空值）
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        white_rule = FormulaRule(
            formula=[f'ISBLANK(${notes_col_letter}2)'],
            fill=white_fill
        )
        self.summary_sheet.conditional_formatting.add(
            f"A2:{openpyxl.utils.get_column_letter(notes_col - 1)}{max_row}",
            white_rule
        )


if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = ExcelDataProcessor()
    ex.show()
    sys.exit(app.exec())