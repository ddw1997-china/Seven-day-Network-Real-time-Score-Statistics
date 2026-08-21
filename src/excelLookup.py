import sys
import os
from PyQt6.QtWidgets import (QApplication, QMainWindow, QVBoxLayout, QWidget, QPushButton,
                             QLabel, QFileDialog, QMessageBox, QTableWidget, QTableWidgetItem,
                             QComboBox, QTabWidget, QHBoxLayout, QGroupBox)
from PyQt6.QtCore import Qt
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string


class XLookupApp(QWidget):
    def __init__(self):
        super().__init__()
        # 存储变量
        self.file_a = ""
        self.file_b = ""
        self.file_a_data = []
        self.file_b_data = []
        self.file_a_headers = []
        self.file_b_headers = []
        self.file_a_sheets = []
        self.file_b_sheets = []

        # 选择的列
        self.lookup_col = None
        self.search_col = None
        self.return_col = None
        self.output_col = None

        # 初始化UI
        self.init_ui()

    def init_ui(self):
        # 主布局
        main_layout = QVBoxLayout()

        # 文件选择部分
        file_group = QGroupBox("文件选择")
        file_group.setObjectName("file_group")
        file_layout = QHBoxLayout()
        file_layout.addStretch()

        self.file_a_btn = QPushButton("加载A文件(源数据)")
        self.file_a_btn.setFixedWidth(120)
        self.file_a_btn.clicked.connect(lambda: self.load_file("a"))
        file_layout.addWidget(self.file_a_btn)
        file_layout.addStretch()

        self.file_b_btn = QPushButton("加载B文件(查找数据)")
        self.file_b_btn.setFixedWidth(120)
        self.file_b_btn.clicked.connect(lambda: self.load_file("b"))
        file_layout.addWidget(self.file_b_btn)
        file_layout.addStretch()

        file_group.setLayout(file_layout)
        main_layout.addWidget(file_group)

        # 表格和列选择部分
        tab_widget = QTabWidget()

        # A文件标签页
        self.a_tab = QWidget()
        a_tab_layout = QVBoxLayout()

        self.a_table = QTableWidget()
        a_tab_layout.addWidget(self.a_table)

        a_col_group = QGroupBox("A文件列选择")
        a_col_group.setObjectName("a_col_group")
        a_col_layout = QHBoxLayout()
        a_col_layout.addStretch()
        self.a_sheet_combo = QComboBox()
        self.a_sheet_combo.setPlaceholderText("选择表格")
        self.a_sheet_combo.currentTextChanged.connect(self.on_a_sheet_changed)
        a_col_layout.addWidget(QLabel("工作表:"))
        a_col_layout.addWidget(self.a_sheet_combo)
        a_col_layout.addStretch()

        self.lookup_combo = QComboBox()
        self.lookup_combo.setPlaceholderText("选择查询列")
        self.lookup_combo.currentTextChanged.connect(self.check_execute_conditions)
        a_col_layout.addWidget(QLabel("查询列:"))
        a_col_layout.addWidget(self.lookup_combo)
        a_col_layout.addStretch()

        self.output_combo = QComboBox()
        self.output_combo.setPlaceholderText("选择结果列")
        self.output_combo.currentTextChanged.connect(self.check_execute_conditions)
        a_col_layout.addWidget(QLabel("结果列:"))
        a_col_layout.addWidget(self.output_combo)
        a_col_layout.addStretch()
        a_col_group.setLayout(a_col_layout)
        a_col_layout.setContentsMargins(0,20,0,0)
        a_tab_layout.addWidget(a_col_group)

        self.a_tab.setLayout(a_tab_layout)
        tab_widget.addTab(self.a_tab, "A文件")

        # B文件标签页
        self.b_tab = QWidget()
        b_tab_layout = QVBoxLayout()

        self.b_table = QTableWidget()
        b_tab_layout.addWidget(self.b_table)

        b_col_group = QGroupBox("B文件列选择")
        b_col_group.setObjectName("b_col_group")
        b_col_layout = QHBoxLayout()
        b_col_layout.addStretch()
        self.b_sheet_combo = QComboBox()
        self.b_sheet_combo.setPlaceholderText("选择表格")
        self.b_sheet_combo.currentTextChanged.connect(self.on_b_sheet_changed)
        b_col_layout.addWidget(QLabel("工作表:"))
        b_col_layout.addWidget(self.b_sheet_combo)
        b_col_layout.addStretch()

        self.search_combo = QComboBox()
        self.search_combo.setPlaceholderText("选择搜索列")
        self.search_combo.currentTextChanged.connect(self.check_execute_conditions)
        b_col_layout.addWidget(QLabel("搜索列:"))
        b_col_layout.addWidget(self.search_combo)
        b_col_layout.addStretch()

        self.return_combo = QComboBox()
        self.return_combo.setPlaceholderText("选择返回列")
        self.return_combo.currentTextChanged.connect(self.check_execute_conditions)
        b_col_layout.addWidget(QLabel("返回列:"))
        b_col_layout.addWidget(self.return_combo)
        b_col_layout.addStretch()
        b_col_group.setLayout(b_col_layout)
        b_col_layout.setContentsMargins(0,20,0,0)
        b_tab_layout.addWidget(b_col_group)

        self.b_tab.setLayout(b_tab_layout)
        tab_widget.addTab(self.b_tab, "B文件")

        main_layout.addWidget(tab_widget)

        # 执行按钮
        self.execute_btn = QPushButton("执行XLOOKUP")
        self.execute_btn.setFixedWidth(120)
        self.execute_btn.setEnabled(False)
        self.execute_btn.clicked.connect(self.execute_xlookup)
        main_layout.addWidget(self.execute_btn,alignment=Qt.AlignmentFlag.AlignHCenter)

        # 状态信息
        self.status_label = QLabel("准备就绪")
        self.status_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        main_layout.addWidget(self.status_label)

        self.setLayout(main_layout)

    def check_execute_conditions(self):
        """检查执行条件并更新按钮状态"""
        self.lookup_col = self.lookup_combo.currentText() if self.lookup_combo.currentIndex() >= 0 else None
        self.search_col = self.search_combo.currentText() if self.search_combo.currentIndex() >= 0 else None
        self.return_col = self.return_combo.currentText() if self.return_combo.currentIndex() >= 0 else None
        self.output_col = self.output_combo.currentText() if self.output_combo.currentIndex() >= 0 else None

        can_execute = all([
            self.file_a,
            self.file_b,
            self.lookup_col,
            self.search_col,
            self.return_col,
            self.output_col
        ])

        self.execute_btn.setEnabled(can_execute)
        self.update_status_label()

    def update_status_label(self):
        """更新状态栏信息"""
        status_parts = []
        if self.file_a:
            status_parts.append(f"A文件: {os.path.basename(self.file_a)}")
            if self.lookup_col:
                status_parts.append(f"查询列: {self.lookup_col}")
            if self.output_col:
                status_parts.append(f"结果列: {self.output_col}")

        if self.file_b:
            status_parts.append(f"B文件: {os.path.basename(self.file_b)}")
            if self.search_col:
                status_parts.append(f"搜索列: {self.search_col}")
            if self.return_col:
                status_parts.append(f"返回列: {self.return_col}")

        self.status_label.setText(" | ".join(status_parts) if status_parts else "准备就绪")

    def load_file(self, file_type):
        """加载Excel文件"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, f"选择{'A' if file_type == 'a' else 'B'}文件", "", "Excel文件 (*.xlsx *.xlsm *.xls)")

        if not file_path:
            return

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            sheets = wb.sheetnames

            if file_type == "a":
                self.file_a = file_path
                self.file_a_sheets = sheets
                self.a_sheet_combo.clear()
                self.a_sheet_combo.addItems(sheets)
                self.load_sheet_data("a", sheets[0] if sheets else "")
            else:
                self.file_b = file_path
                self.file_b_sheets = sheets
                self.b_sheet_combo.clear()
                self.b_sheet_combo.addItems(sheets)
                self.load_sheet_data("b", sheets[0] if sheets else "")

            wb.close()
            self.check_execute_conditions()

        except Exception as e:
            QMessageBox.critical(self, "错误", f"无法加载文件: {str(e)}")

    def load_sheet_data(self, file_type, sheet_name):
        """加载工作表数据"""
        if not sheet_name:
            return

        file_path = self.file_a if file_type == "a" else self.file_b
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb[sheet_name]

            headers = []
            data = []

            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:  # 第一行作为表头
                    headers = [str(h) if h is not None else f"列{get_column_letter(j + 1)}"
                               for j, h in enumerate(row)]
                else:
                    data.append(row)

            if file_type == "a":
                self.file_a_data = data
                self.file_a_headers = headers
                self.update_table(self.a_table, headers, data)
                self.update_column_combos()
            else:
                self.file_b_data = data
                self.file_b_headers = headers
                self.update_table(self.b_table, headers, data)
                self.update_column_combos()

            wb.close()
            self.check_execute_conditions()

        except Exception as e:
            QMessageBox.critical(self, "错误", f"无法加载工作表数据: {str(e)}")

    def update_table(self, table, headers, data):
        """更新表格显示"""
        table.clear()
        table.setColumnCount(len(headers))
        table.setRowCount(len(data))
        table.setHorizontalHeaderLabels(headers)

        for i, row in enumerate(data):
            for j, value in enumerate(row):
                item = QTableWidgetItem(str(value) if value is not None else "")
                table.setItem(i, j, item)

        table.resizeColumnsToContents()

    def update_column_combos(self):
        """更新列选择下拉框"""
        self.lookup_combo.clear()
        self.output_combo.clear()
        if self.file_a_headers:
            self.lookup_combo.addItems(self.file_a_headers)
            self.output_combo.addItems(self.file_a_headers)

        self.search_combo.clear()
        self.return_combo.clear()
        if self.file_b_headers:
            self.search_combo.addItems(self.file_b_headers)
            self.return_combo.addItems(self.file_b_headers)

        self.check_execute_conditions()

    def on_a_sheet_changed(self, sheet_name):
        """A文件工作表切换"""
        if self.file_a and sheet_name:
            self.load_sheet_data("a", sheet_name)

    def on_b_sheet_changed(self, sheet_name):
        """B文件工作表切换"""
        if self.file_b and sheet_name:
            self.load_sheet_data("b", sheet_name)

    def execute_xlookup(self):
        """执行XLOOKUP操作"""
        try:
            # 获取列索引
            lookup_idx = self.file_a_headers.index(self.lookup_col)
            search_idx = self.file_b_headers.index(self.search_col)
            return_idx = self.file_b_headers.index(self.return_col)
            output_idx = self.file_a_headers.index(self.output_col)

            # 创建搜索字典（过滤B文件中的空值）
            search_dict = {}
            for row in self.file_b_data:
                search_key = row[search_idx]
                if search_key is not None and str(search_key).strip() != "":
                    search_dict[search_key] = row[return_idx]
                    print(search_key, search_dict[search_key])

            # 打开A文件写入结果
            wb = openpyxl.load_workbook(self.file_a)
            ws = wb[self.a_sheet_combo.currentText()]

            updated_count = 0
            skipped_count = 0

            for i, row in enumerate(self.file_a_data, start=2):  # 从第2行开始
                lookup_val = row[lookup_idx]

                # 跳过空值查询
                if lookup_val is None or str(lookup_val).strip() == "":
                    skipped_count += 1
                    continue

                result_val = search_dict.get(lookup_val, "#N/A")
                ws.cell(row=i, column=output_idx + 1, value=result_val)
                updated_count += 1

            # 保存文件
            wb.save(self.file_a)
            wb.close()

            # 显示结果
            QMessageBox.information(self, "完成",
                                    f"成功处理 {updated_count} 行数据\n"
                                    f"跳过 {skipped_count} 个空值查询\n"
                                    f"结果已写入A文件的 {self.output_col} 列")

            # 重新加载A文件显示更新后的数据
            self.load_sheet_data("a", self.a_sheet_combo.currentText())

        except Exception as e:
            QMessageBox.critical(self, "错误", f"执行XLOOKUP时出错: {str(e)}")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = XLookupApp()
    window.show()
    sys.exit(app.exec())