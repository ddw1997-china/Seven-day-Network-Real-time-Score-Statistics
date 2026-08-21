import sys
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
                             QLabel, QLineEdit, QPushButton, QMessageBox, QFileDialog, QCheckBox, QGridLayout)
from PyQt6.QtCore import Qt
from openpyxl import load_workbook
import os
import subprocess


class ExcelFiller(QWidget):
    def __init__(self):
        super().__init__()
        self.setup_ui()

    def setup_ui(self):
        # 主窗口部件
        main_layout = QVBoxLayout(self)

        main_layout.addStretch(1)
        main_layout.setSpacing(20)
        grid_h_layout = QHBoxLayout()
        grid_layout = QGridLayout()
        grid_layout.setVerticalSpacing(20)
        grid_h_layout.addStretch()
        grid_h_layout.addLayout(grid_layout)
        grid_h_layout.addStretch()
        main_layout.addLayout(grid_h_layout)

        # 输入字段
        self.create_input_field(grid_layout,0, "开始填充的行号（默认2）:", "entry_row")
        self.create_input_field(grid_layout,1, "重复填充的序列（如a，b，c用逗号分隔）:", "entry_abcd")
        self.create_input_field(grid_layout,2, "序列中每项（如a，b，c）的值（如2,3,6用逗号分隔）:", "entry_1234")
        self.create_input_field(grid_layout,3, "填充的总行数参考哪列（从1开始）:", "entry_total_column")
        self.create_input_field(grid_layout,4, "填充到哪列（从1开始）:", "entry_write_column")

        # 选项行 - 使用水平布局
        options_layout = QHBoxLayout()
        options_layout.setSpacing(20)
        options_layout.addStretch()

        # 场次选项
        self.checkbox_session = QCheckBox("添加场次信息")
        self.checkbox_session.setChecked(False)  # 默认选中
        options_layout.addWidget(self.checkbox_session,alignment=Qt.AlignmentFlag.AlignBottom | Qt.AlignmentFlag.AlignHCenter)

        # 座位选项
        self.checkbox_seat = QCheckBox("添加座位号")
        self.checkbox_seat.setChecked(False)  # 默认选中
        options_layout.addWidget(self.checkbox_seat,alignment=Qt.AlignmentFlag.AlignBottom | Qt.AlignmentFlag.AlignHCenter)
        options_layout.addStretch()
        # 将选项行添加到主布局
        main_layout.addLayout(options_layout)

        # 提交按钮
        submit_button = QPushButton("选择文件并生成Excel")
        submit_button.setFixedWidth(200)
        submit_button.clicked.connect(self.on_submit)
        main_layout.addWidget(submit_button,alignment=Qt.AlignmentFlag.AlignBottom | Qt.AlignmentFlag.AlignHCenter)
        main_layout.addStretch(2)
    def create_input_field(self, layout,row, label_text, field_name):
        label = QLabel(label_text)
        label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)  # 右对齐
        line_edit = QLineEdit()
        if row == 0:
            line_edit.setText("2")
        else:
            line_edit.setPlaceholderText("必填")

        line_edit.setFixedWidth(150)
        layout.addWidget(label, row, 2)  # 第2行，跨越0-1列
        layout.addWidget(line_edit, row, 3)

        setattr(self, field_name, line_edit)  # 动态设置属性

    def fill_excel(self, file_path, ab_values, mn_values, total_column, write_column, save_path):
        try:
            # 加载Excel文件
            wb = load_workbook(file_path)
            ws = wb.active

            # 获取总行数列的总行数
            total_rows = 0
            for row in ws.iter_rows(min_col=total_column, max_col=total_column, values_only=True):
                if row[0] is not None:  # 忽略空单元格
                    total_rows += 1
            print(f"总行数: {total_rows}")  # 调试信息

            # 检查输入的数组长度是否一致
            if len(ab_values) != len(mn_values):
                QMessageBox.critical(self, "错误", "a/b的数组长度和m/n的数组长度不一致")
                return

            # 获取开始行号
            start_index = int(self.entry_row.text())
            row_index = start_index  # 从用户指定的行开始

            mn__ = [int(strValue) for strValue in mn_values]
            __mn = sum(mn__)
            while row_index <= total_rows:
                for ab, mn in zip(ab_values, mn_values):
                    print(f"当前行: {row_index}, ab: {ab}, mn: {mn}")  # 调试信息
                    for _mn in range(int(mn)):
                        if row_index > total_rows:
                            break
                        # 将数据写入到用户选择的写入列中
                        ws.cell(row=row_index, column=write_column, value=ab)

                        # 如果勾选了场次信息
                        if self.checkbox_session.isChecked():
                            num = 1 + (row_index - start_index) // int(__mn)
                            ws.cell(row=row_index, column=write_column + 1, value=f"第{num}场")

                        # 如果勾选了座位号
                        if self.checkbox_seat.isChecked():
                            # seat_num = 1 + (row_index - start_index) % int(mn)
                            seat_num = 1 + _mn
                            col_offset = 1 if self.checkbox_session.isChecked() else 0
                            ws.cell(row=row_index, column=write_column + 1 + col_offset, value=f"{seat_num}")

                        print(f"写入行: {row_index}, 列: {write_column}, 值: {ab}")  # 调试信息
                        row_index += 1

            # 另存为新文件
            wb.save(save_path)
            QMessageBox.information(self, "完成", f"数据已成功填充并保存到 {save_path}")

            # 自动打开文件
            if os.name == 'nt':  # Windows
                os.startfile(save_path)
            else:  # macOS 或 Linux
                subprocess.run(['open', save_path] if os.name == 'posix' else ['xdg-open', save_path])
        except Exception as e:
            QMessageBox.critical(self, "错误", f"发生错误: {str(e)}")

    def replace_chinese_comma(self, text):
        """将字符串中的中文逗号替换为英文逗号"""
        return text.replace("，", ",")

    def on_submit(self):
        try:
            # 获取用户输入的字符串并替换中文逗号
            ab_values = self.replace_chinese_comma(self.entry_abcd.text().strip()).split(',')
            mn_values = self.replace_chinese_comma(self.entry_1234.text().strip()).split(',')

            # 检查输入是否为空
            if not ab_values or not mn_values:
                QMessageBox.critical(self, "错误", "输入不能为空")
                return

            # 选择Excel文件
            file_path, _ = QFileDialog.getOpenFileName(self, "选择Excel文件", "", "Excel文件 (*.xlsx)")
            if not file_path:
                return

            # 获取列号
            total_column = int(self.entry_total_column.text())
            write_column = int(self.entry_write_column.text())
            if total_column < 1 or write_column < 1:
                QMessageBox.critical(self, "错误", "列号必须大于0")
                return

            # 获取开始行号
            start_index = int(self.entry_row.text())
            if start_index < 1:
                QMessageBox.critical(self, "错误", "开始行号必须大于0")
                return

            # 选择保存路径
            save_path, _ = QFileDialog.getSaveFileName(self, "保存文件", "", "Excel文件 (*.xlsx)")
            if save_path:
                self.fill_excel(file_path, ab_values, mn_values, total_column, write_column, save_path)
        except ValueError:
            QMessageBox.critical(self, "错误", "请输入有效的整数")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = ExcelFiller()
    window.show()
    sys.exit(app.exec())